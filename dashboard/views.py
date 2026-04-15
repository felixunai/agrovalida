import csv
import io
from django.shortcuts import render
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q
from django.utils import timezone
from django.conf import settings
from lotes.models import Lote
from defensivos.models import Defensivo, ClasseDefensivo
from notas.models import NotaFiscal
from accounts.models import Plano
from accounts.decorators import plano_requerido


@login_required
def dashboard(request):
    hoje = timezone.now().date()
    dias_alerta = getattr(settings, 'ALERTA_DIAS_VENCIMENTO', 90)
    limite_alerta = hoje + timezone.timedelta(days=dias_alerta)

    # Apenas lotes ativos no painel
    user_lotes = Lote.objects.filter(cadastrado_por=request.user, ativo=True)

    lotes_vencidos = user_lotes.filter(data_validade__lt=hoje).select_related('defensivo', 'fazenda')
    lotes_vencendo = user_lotes.filter(
        data_validade__gte=hoje, data_validade__lte=limite_alerta
    ).select_related('defensivo', 'fazenda')
    lotes_vigentes = user_lotes.filter(data_validade__gt=limite_alerta).select_related('defensivo', 'fazenda')

    total_defensivos = Defensivo.objects.filter(cadastrado_por=request.user, ativo=True).count()
    total_lotes = user_lotes.count()
    total_vencidos = lotes_vencidos.count()
    total_vencendo = lotes_vencendo.count()
    total_vigentes = lotes_vigentes.count()
    total_notas = NotaFiscal.objects.filter(cadastrado_por=request.user).count()
    notas_pendentes = NotaFiscal.objects.filter(cadastrado_por=request.user, processado=False).count()

    lotes_proximos = (lotes_vencendo | lotes_vencidos).order_by('data_validade')[:20]

    por_classe = user_lotes.values(
        'defensivo__classe'
    ).annotate(
        total=Count('id'),
        vencidos=Count('id', filter=Q(data_validade__lt=hoje)),
        vencendo=Count('id', filter=Q(
            data_validade__gte=hoje, data_validade__lte=limite_alerta
        )),
    ).order_by('-total')

    plano_pro = Plano.objects.filter(ativo=True, preco__gt=0).order_by('preco').first()

    context = {
        'total_defensivos': total_defensivos,
        'total_lotes': total_lotes,
        'total_vencidos': total_vencidos,
        'total_vencendo': total_vencendo,
        'total_vigentes': total_vigentes,
        'total_notas': total_notas,
        'notas_pendentes': notas_pendentes,
        'lotes_proximos': lotes_proximos,
        'por_classe': por_classe,
        'classes': ClasseDefensivo.choices,
        'dias_alerta': dias_alerta,
        'plano_pro': plano_pro,
    }
    return render(request, 'dashboard/index.html', context)


def _filtrar_lotes_relatorio(user, params):
    """Shared filtering logic for the expiry report (HTML, CSV and Excel views)."""
    hoje = timezone.now().date()
    dias_alerta = getattr(settings, 'ALERTA_DIAS_VENCIMENTO', 90)
    limite_alerta = hoje + timezone.timedelta(days=dias_alerta)

    status              = params.get('status', 'todos')
    classe              = params.get('classe', '')
    fornecedor          = params.get('fornecedor', '')
    local_armazenamento = params.get('local_armazenamento', '')
    fazenda_filtro      = params.get('fazenda', '')

    lotes = (
        Lote.objects
        .filter(cadastrado_por=user)
        .select_related('defensivo', 'fazenda')
        .order_by('data_validade')
    )

    if status == 'vencido':
        lotes = lotes.filter(data_validade__lt=hoje)
    elif status == 'vencendo':
        lotes = lotes.filter(data_validade__gte=hoje, data_validade__lte=limite_alerta)
    elif status == 'vigente':
        lotes = lotes.filter(data_validade__gt=limite_alerta)

    if classe:
        lotes = lotes.filter(defensivo__classe=classe)
    if fornecedor:
        lotes = lotes.filter(fornecedor__icontains=fornecedor)
    if local_armazenamento:
        lotes = lotes.filter(local_armazenamento__icontains=local_armazenamento)
    if fazenda_filtro:
        lotes = lotes.filter(fazenda__nome__icontains=fazenda_filtro)

    return lotes, status, classe, fornecedor, local_armazenamento, fazenda_filtro, dias_alerta


@login_required
@plano_requerido
def relatorio_vencimento(request):
    lotes, status, classe, fornecedor, local_armazenamento, fazenda_filtro, dias_alerta = _filtrar_lotes_relatorio(request.user, request.GET)

    from fazendas.models import Fazenda
    fazendas = Fazenda.objects.filter(cadastrado_por=request.user).order_by('nome')

    context = {
        'lotes': lotes,
        'total_count': lotes.count(),
        'status': status,
        'classe': classe,
        'fornecedor': fornecedor,
        'local_armazenamento': local_armazenamento,
        'fazenda_filtro': fazenda_filtro,
        'fazendas': fazendas,
        'classes': ClasseDefensivo.choices,
        'dias_alerta': dias_alerta,
    }
    return render(request, 'dashboard/relatorio.html', context)


_STATUS_LABELS = {'vencido': 'Vencido', 'vencendo': 'Vencendo em breve', 'vigente': 'Vigente'}

_RELATORIO_HEADERS = [
    'Nº Lote', 'Produto', 'Classe', 'Fabricação', 'Validade',
    'Dias para Vencer', 'Status', 'Quantidade', 'Unidade',
    'Fornecedor', 'Nota Fiscal', 'Fazenda', 'Local Armazenamento',
]


def _lote_row(lote):
    return [
        lote.numero_lote,
        lote.defensivo.nome_comercial,
        lote.defensivo.get_classe_display(),
        lote.data_fabricacao.strftime('%d/%m/%Y') if lote.data_fabricacao else '',
        lote.data_validade.strftime('%d/%m/%Y'),
        lote.dias_para_vencer,
        _STATUS_LABELS.get(lote.status_vencimento, lote.status_vencimento),
        str(lote.quantidade).replace('.', ','),
        lote.get_unidade_display(),
        lote.fornecedor,
        lote.nota_fiscal,
        lote.fazenda.nome if lote.fazenda else '',
        lote.local_armazenamento,
    ]


@login_required
@plano_requerido
def relatorio_vencimento_csv(request):
    lotes, *_ = _filtrar_lotes_relatorio(request.user, request.GET)

    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="relatorio_vencimento.csv"'

    writer = csv.writer(response, delimiter=';')
    writer.writerow(_RELATORIO_HEADERS)
    for lote in lotes:
        writer.writerow(_lote_row(lote))

    return response


@login_required
@plano_requerido
def relatorio_vencimento_excel(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse('openpyxl não instalado.', status=500)

    lotes, *_ = _filtrar_lotes_relatorio(request.user, request.GET)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Relatório de Vencimento'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='1a3a2a')

    for col_idx, header in enumerate(_RELATORIO_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    FILL_VENCIDO  = PatternFill(fill_type='solid', fgColor='FFDDE2')
    FILL_VENCENDO = PatternFill(fill_type='solid', fgColor='FFF3CD')

    for row_idx, lote in enumerate(lotes, start=2):
        row_data = _lote_row(lote)
        # Store dias_para_vencer as int for proper Excel number
        row_data[5] = lote.dias_para_vencer
        row_data[7] = float(lote.quantidade)
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            status = lote.status_vencimento
            if status == 'vencido':
                cell.fill = FILL_VENCIDO
            elif status == 'vencendo':
                cell.fill = FILL_VENCENDO

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="relatorio_vencimento.xlsx"'
    return response


@login_required
@plano_requerido
def dashboard_graficos(request):
    import json
    hoje = timezone.now().date()
    dias_alerta = getattr(settings, 'ALERTA_DIAS_VENCIMENTO', 90)
    limite_alerta = hoje + timezone.timedelta(days=dias_alerta)

    user_lotes = Lote.objects.filter(cadastrado_por=request.user, ativo=True)

    # 1. Lotes por status
    total_vencidos  = user_lotes.filter(data_validade__lt=hoje).count()
    total_vencendo  = user_lotes.filter(data_validade__gte=hoje, data_validade__lte=limite_alerta).count()
    total_vigentes  = user_lotes.filter(data_validade__gt=limite_alerta).count()

    # 2. Lotes por classe
    por_classe_qs = (
        user_lotes
        .values('defensivo__classe')
        .annotate(total=Count('id'))
        .order_by('-total')
    )
    classes_dict = dict(ClasseDefensivo.choices)
    classe_labels = [classes_dict.get(item['defensivo__classe'], item['defensivo__classe']) for item in por_classe_qs]
    classe_values = [item['total'] for item in por_classe_qs]

    # 3. Vencimentos por mês (próximos 12 meses)
    from collections import defaultdict
    import calendar
    meses_labels = []
    meses_vencendo = []
    meses_vencido = []
    for i in range(12):
        ano = (hoje.replace(day=1) + timezone.timedelta(days=32 * i)).year
        mes = (hoje.replace(day=1) + timezone.timedelta(days=32 * i)).month
        ultimo_dia = calendar.monthrange(ano, mes)[1]
        inicio = hoje.__class__(ano, mes, 1)
        fim    = hoje.__class__(ano, mes, ultimo_dia)
        count  = user_lotes.filter(data_validade__gte=inicio, data_validade__lte=fim).count()
        meses_labels.append(f'{mes:02d}/{ano}')
        if fim < hoje:
            meses_vencido.append(count)
            meses_vencendo.append(0)
        else:
            meses_vencido.append(0)
            meses_vencendo.append(count)

    # 4. Lotes por fazenda
    por_fazenda_qs = (
        user_lotes
        .values('fazenda__nome')
        .annotate(total=Count('id'))
        .order_by('-total')
    )
    fazenda_labels = [item['fazenda__nome'] or 'Sem fazenda' for item in por_fazenda_qs]
    fazenda_values = [item['total'] for item in por_fazenda_qs]

    # 5. Produtos mais estocados (top 8)
    por_produto_qs = (
        user_lotes
        .values('defensivo__nome_comercial')
        .annotate(total=Count('id'))
        .order_by('-total')[:8]
    )
    produto_labels = [item['defensivo__nome_comercial'] for item in por_produto_qs]
    produto_values = [item['total'] for item in por_produto_qs]

    context = {
        'status_data': json.dumps({
            'labels': ['Vigentes', 'Vencendo em breve', 'Vencidos'],
            'values': [total_vigentes, total_vencendo, total_vencidos],
        }),
        'classe_data': json.dumps({'labels': classe_labels, 'values': classe_values}),
        'meses_data': json.dumps({
            'labels': meses_labels,
            'vencendo': meses_vencendo,
            'vencido': meses_vencido,
        }),
        'fazenda_data': json.dumps({'labels': fazenda_labels, 'values': fazenda_values}),
        'produto_data': json.dumps({'labels': produto_labels, 'values': produto_values}),
        'total_lotes': user_lotes.count(),
    }
    return render(request, 'dashboard/graficos.html', context)
